import pytesseract
#
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# print(pytesseract.image_to_string(r'D:\pythonProject\just_dial\contact\ 1st May Advertisement.png'))
import os
from os import listdir,path
from os.path import isfile
from openpyxl import load_workbook
wb = load_workbook(r"D:\pythonProject\just_dial\data_full 2.xlsx")
ws = wb.active

path_s = "D://pythonProject//just_dial//contact//"
dir_list = os.listdir(path=path_s)
# print(isfile(path_s))
# for ds in dir_list:
#     print(ds[1:-4])
image_path = r"D:\pythonProject\just_dial\contact\ "
for r in range(1, ws.max_row+1):
    # print(ws.cell(row=r,column=1).value)
    for ds in dir_list:
        # print(ds[1:-4])
        # print(ws.cell(row=r,column=1).value)
        if ds[1:-4] == ws.cell(row=r, column=1).value:

            print(image_path[:-1] + ds)
            # print(pytesseract.image_to_string(r'D:\pythonProject\just_dial\contact\ Red Carpet Ad Agencies.png'))
            text_value = pytesseract.image_to_string(image_path[:-1] + ds)
            print(text_value)
            ws.cell(row=r,column=5).value = text_value
            wb.save(r"D:\pythonProject\just_dial\data_full 3.xlsx")
