from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.common.alert import Alert
driver = webdriver.Chrome(r"D:\pythonProject\just_dial\driver\chromedriver.exe")

# url = 'https://www.justdial.com/Chennai/Ad-Film-Makers/nct-10890946'
# driver.get(url)

wb = Workbook()
ws = wb.active
driver.implicitly_wait(30)
# Alert(driver).dismiss()
# page-44
l = 0
for r in range(1,45):
    print(r)
    driver.get(url='https://www.justdial.com/Chennai/Ad-Film-Makers/nct-10890946/page-'+str(r))
    driver.implicitly_wait(30)
    for heading in driver.find_elements_by_class_name("rsl"):
        for cont in heading.find_elements_by_class_name("store-details"):
            l = l +1
            ws.cell(row=l,column=5).value = 6
            # print(cont.text)
            title = cont.find_element_by_class_name('store-name').text
            print(title)
            ws.cell(row=l,column=1).value = title
            try:
                image = cont.find_element_by_class_name("contact-info ").screenshot(r"D:\pythonProject\just_dial\contact\ "+title+".png")
            except:
                pass
            for list_rating in cont.find_elements_by_class_name("newrtings "):
                print(list_rating.find_element_by_class_name('green-box').text)
                ws.cell(row=l, column=2).value = list_rating.find_element_by_class_name('green-box').text
                print(list_rating.find_element_by_class_name('rt_count').text)
                ws.cell(row=l, column=3).value = list_rating.find_element_by_class_name('rt_count').text
                print(list_rating.find_element_by_tag_name('a').get_attribute('href'))
                ws.cell(row=l, column=4).value = list_rating.find_element_by_tag_name('a').get_attribute('href')
                # ws.cell(row=l, column=5).value = heading.find_element_by_class_name("contact-info ").screenshot(r"D:\pythonProject\just_dial\contact\ "+title+".png")
                wb.save(r"D:\pythonProject\just_dial\data_full 2.xlsx")
