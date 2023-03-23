import xlsxwriter
from openpyxl import load_workbook

# wb = load_workbook(r"D:\pythonProject\just_dial\data.xlsx")
workbook = xlsxwriter.Workbook('images.xlsx')
worksheet = workbook.add_worksheet()
worksheet.insert_image('B2', 'D:\pythonProject\just_dial\contact\ 1st May Advertisement.png')
workbook.close()